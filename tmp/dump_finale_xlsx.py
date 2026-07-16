from pathlib import Path
import openpyxl

p = Path(r"database/input/clubmeisterschaft_donaubowler/Clubpokal DB 2026 Finale.xlsx")
wb = openpyxl.load_workbook(p, data_only=True)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name!r} dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} ===")
    for r in range(1, min(ws.max_row or 1, 100) + 1):
        vals = []
        for c in range(1, min(ws.max_column or 1, 25) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(f"{c}:{v!r}")
        if vals:
            print(f"R{r}: {' | '.join(vals)}")
