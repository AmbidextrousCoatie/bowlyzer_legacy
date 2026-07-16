"""Inspect qualifying workbook layout for handicap cols and set blocks."""
from pathlib import Path
import openpyxl

p = Path(r"database/input/clubmeisterschaft_donaubowler/Clubpokal DB 2026.xlsx")
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb.active
print("title R1B:", ws.cell(1, 2).value)
print("max_col", ws.max_column)

# dump row 2 fully
print("\nRow2 all non-empty:")
for c in range(1, (ws.max_column or 1) + 1):
    v = ws.cell(2, c).value
    if v is not None:
        print(f"  col {c}: {v!r}")

print("\nRow1 all non-empty:")
for c in range(1, (ws.max_column or 1) + 1):
    v = ws.cell(1, c).value
    if v is not None:
        print(f"  col {c}: {v!r}")

# row 3 (Hartfeil) — all values
print("\nRow3 (leader) all non-empty:")
for c in range(1, (ws.max_column or 1) + 1):
    v = ws.cell(3, c).value
    if v is not None:
        print(f"  col {c}: {v!r}")

# Check set block pattern: 4 games + HDC + Summe = 6 cols starting at 3
print("\nSet blocks inferred from headers:")
for c in range(1, 43):
    v = ws.cell(2, c).value
    if v:
        print(f"  {c}: {v}")
