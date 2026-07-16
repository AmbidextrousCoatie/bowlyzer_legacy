"""Analyze Clubmeisterschaft qualifying + finals workbooks."""
from __future__ import annotations

from pathlib import Path
from collections import defaultdict
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
IN_DIR = ROOT / "database" / "input" / "clubmeisterschaft_donaubowler"

FINALE = IN_DIR / "Clubpokal DB 2026 Finale.xlsx"
QUAL = IN_DIR / "Clubpokal DB 2026.xlsx"


def dump_sheet(path: Path, max_rows: int = 40, max_cols: int = 45) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"\n#### {path.name} sheet={wb.sheetnames[0]!r} {ws.max_row}x{ws.max_column}")
    for r in range(1, min(ws.max_row or 1, max_rows) + 1):
        vals = []
        for c in range(1, min(ws.max_column or 1, max_cols) + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                vals.append(f"{c}:{v!r}")
        if vals:
            print(f"R{r}: {' | '.join(vals)}")


def analyze_finale() -> None:
    wb = openpyxl.load_workbook(FINALE, data_only=True)
    ws = wb.active
    # headers row 2
    headers = {c: ws.cell(2, c).value for c in range(1, 16)}
    print("\n==== FINALE STRUCTURE ====")
    print("headers:", headers)
    print("row1:", {c: ws.cell(1, c).value for c in range(1, 16) if ws.cell(1, c).value})

    # cols: 1 seed, 2 name, 3 hdc, 4-9 scratch games, 10-15 inkl hdc
    scratch_cols = {
        "elim_1": 4,
        "elim_2": 5,
        "sl1": 6,
        "sl2": 7,
        "f1": 8,
        "f2": 9,
    }
    hdc_cols = {
        "elim_1": 10,
        "elim_2": 11,
        "sl1": 12,
        "sl2": 13,
        "f1": 14,
        "f2": 15,
    }

    players = []
    for r in range(3, 9):
        seed = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        hdc = ws.cell(r, 3).value
        if not name:
            continue
        scratch = {k: ws.cell(r, c).value for k, c in scratch_cols.items()}
        inkl = {k: ws.cell(r, c).value for k, c in hdc_cols.items()}
        # verify inkl = scratch + hdc where both present
        mismatches = []
        for k, sv in scratch.items():
            iv = inkl[k]
            if sv is not None and iv is not None and hdc is not None:
                if int(iv) != int(sv) + int(hdc):
                    mismatches.append((k, sv, hdc, iv))
        players.append(
            {
                "seed": int(seed),
                "name": str(name),
                "hdc": int(hdc) if hdc is not None else None,
                "scratch": scratch,
                "inkl": inkl,
                "mismatches": mismatches,
            }
        )

    print("\nPlayers / games:")
    for p in players:
        print(f"  seed {p['seed']}: {p['name']}  HDC={p['hdc']}")
        print(f"    scratch: {p['scratch']}")
        print(f"    inklhdc: {p['inkl']}")
        if p["mismatches"]:
            print(f"    MISMATCH scratch+hdc != inkl: {p['mismatches']}")

    # Elim field: seeds 4-6, 2 games
    elim = [p for p in players if p["seed"] in (4, 5, 6)]
    print("\n---- Elim (seeds 4-6) scratch totals ----")
    for p in elim:
        g1, g2 = p["scratch"]["elim_1"], p["scratch"]["elim_2"]
        games = [x for x in (g1, g2) if x is not None]
        total = sum(int(x) for x in games)
        print(f"  {p['name']}: games={games} total={total} n={len(games)}")

    print("\n---- Elim inkl-hdc totals ----")
    for p in elim:
        g1, g2 = p["inkl"]["elim_1"], p["inkl"]["elim_2"]
        games = [x for x in (g1, g2) if x is not None]
        total = sum(int(x) for x in games)
        print(f"  {p['name']}: games={games} total={total} n={len(games)}")

    print("\n---- Stepladder 1 (seed 3 vs elim winner candidate) ----")
    for p in players:
        if p["scratch"]["sl1"] is not None:
            print(f"  {p['name']} seed={p['seed']}: scratch={p['scratch']['sl1']} inkl={p['inkl']['sl1']}")

    print("\n---- Stepladder 2 ----")
    for p in players:
        if p["scratch"]["sl2"] is not None:
            print(f"  {p['name']} seed={p['seed']}: scratch={p['scratch']['sl2']} inkl={p['inkl']['sl2']}")

    print("\n---- Finals (up to 2 games shown; BO3 may stop early) ----")
    for p in players:
        if p["scratch"]["f1"] is not None or p["scratch"]["f2"] is not None:
            s = [p["scratch"]["f1"], p["scratch"]["f2"]]
            h = [p["inkl"]["f1"], p["inkl"]["f2"]]
            print(f"  {p['name']} seed={p['seed']}: scratch={s} inkl={h}")
            # game wins scratch
            others = [q for q in players if q is not p and (q["scratch"]["f1"] is not None)]
            if others:
                o = others[0]
                wins = 0
                for a, b in zip(s, [o["scratch"]["f1"], o["scratch"]["f2"]]):
                    if a is not None and b is not None and a > b:
                        wins += 1
                print(f"    vs {o['name']}: scratch game-wins (partial)={wins}")

    print("\n---- Implied path ----")
    print("  #1 Feller bye to final")
    print("  Elim participants: Schneider(4), Rettinger(5), Luu(6)")
    print("  SL1: Obermeier(3) vs Schneider — Schneider higher")
    print("  SL2: Harteil(2) vs Schneider — Schneider higher")
    print("  Final: Feller vs Schneider — Feller wins both games (2-0 BO3)")


def summarize_qual() -> None:
    wb = openpyxl.load_workbook(QUAL, data_only=True)
    ws = wb.active
    print("\n==== QUALIFYING Clubpokal DB 2026.xlsx ====")
    print(f"sheet={wb.sheetnames[0]} rows={ws.max_row} cols={ws.max_column}")
    # count players with names in col B
    names = []
    for r in range(2, (ws.max_row or 1) + 1):
        n = ws.cell(r, 2).value
        if n and str(n).strip():
            names.append(str(n).strip())
    print(f"named players: {len(names)}")
    for n in names:
        print(f"  - {n}")
    # set1 first game col C=3
    print("\nheader row1 sample (first 20 cols):")
    print({c: ws.cell(1, c).value for c in range(1, 21) if ws.cell(1, c).value})
    print("AO/AP headers:", ws.cell(1, 41).value, ws.cell(1, 42).value)
    print("AO/AP row2:", ws.cell(2, 41).value, ws.cell(2, 42).value)


if __name__ == "__main__":
    dump_sheet(QUAL, max_rows=25, max_cols=20)
    summarize_qual()
    analyze_finale()
