"""
Import Clubmeisterschaft Donaubowler XLSX into tournament postprocessed CSV.

(Clubpokal is a separate team KO competition over multiple months — not this importer.)

Workbook layout (Clubmeisterschaft 2026 export):
- Col A: rank (ignored for import)
- Col B: player name
- Cols C onward: repeating blocks of 6 columns per set — 4 scratch games, handicap (sheet), total (sheet).
  We only emit scratch scores; handicap column uses computed per-game handicap (see below).
- Col AO (41): reference score for handicap
- Col AP (42): a priori average

Handicap per game: max(0, 0.7 * (reference_score - a_priori_average))  [pins, rounded half up to nearest int]

Player IDs: resolved from merged league CSV (Player / Player ID) with robust name matching.

Output:
- database/data/tournament_clubmeisterschaft_donaubowler_2026_postprocessed.csv (batch snapshot)
- Merges into database/data/tournament_manual_postprocessed.csv (non–GF-pipeline club/regional imports).
- Removes this event from database/input/gf_tables_export/gf_tournaments_2026__combined_postprocessed.csv if
  it was previously written there, so that file stays GF regional export + Bayerische-only.
- Rebuilds database/data/player_stats_merged_plus_tournaments.csv (league + GF tournaments + manual tournaments).
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import math
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_INPUT_DIR = ROOT / "database" / "input" / "clubmeisterschaft_donaubowler"
DEFAULT_OUTPUT = ROOT / "database" / "data" / "tournament_clubmeisterschaft_donaubowler_2026_postprocessed.csv"
DEFAULT_LEAGUE_CSV = ROOT / "database" / "data" / "league_results_merged.csv"
GF_REGIONAL_TOURNAMENT_CSV = ROOT / "database" / "input" / "gf_tables_export" / "gf_tournaments_2026__combined_postprocessed.csv"
MANUAL_TOURNAMENT_CSV = ROOT / "database" / "data" / "tournament_manual_postprocessed.csv"

COL_NAME = 2
COL_FIRST_SET = 3
COL_REF_HANDICAP = 41  # AO
COL_AP_AVG = 42  # AP
SET_BLOCK_WIDTH = 6
GAMES_PER_SET = 4
NUM_SET_BLOCKS = 6  # six sets in stage 1

EVENT_NAME = "Clubmeisterschaft Donaubowler 2026"


def _rebuild_player_hybrid_local() -> None:
    """League + GF regional tournament export + manual tournament imports (matches database_config)."""
    merged_league = ROOT / "database" / "data" / "league_results_merged.csv"
    out_path = ROOT / "database" / "data" / "player_stats_merged_plus_tournaments.csv"

    league_rows: List[Dict[str, str]] = []
    if merged_league.is_file():
        with merged_league.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            league_rows = [{str(k): str(v or "") for k, v in row.items()} for row in reader]

    tournament_rows: List[Dict[str, str]] = []
    if GF_REGIONAL_TOURNAMENT_CSV.is_file():
        with GF_REGIONAL_TOURNAMENT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            tournament_rows.extend([{str(k): str(v or "") for k, v in row.items()} for row in reader])
    if MANUAL_TOURNAMENT_CSV.is_file():
        with MANUAL_TOURNAMENT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            tournament_rows.extend([{str(k): str(v or "") for k, v in row.items()} for row in reader])

    headers = sorted({k for r in (league_rows + tournament_rows) for k in r.keys()})
    if not headers:
        from database.conversion.bowlingbayern_legacy_core import OUTPUT_HEADERS

        headers = list(OUTPUT_HEADERS)

    out_rows: List[Dict[str, str]] = []
    for row in league_rows:
        merged = {h: str(row.get(h, "")) for h in headers}
        if not str(merged.get("Event Type", "")).strip():
            merged["Event Type"] = "league"
        if not str(merged.get("Event Name", "")).strip():
            merged["Event Name"] = str(merged.get("League", "")).strip()
        if not str(merged.get("Club", "")).strip():
            merged["Club"] = str(merged.get("Team", "")).strip()
        out_rows.append(merged)

    for row in tournament_rows:
        merged = {h: str(row.get(h, "")) for h in headers}
        merged["Input Data"] = "True"
        merged["Computed Data"] = "False"
        out_rows.append(merged)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(out_rows)


def _strip_event_from_path(bmi, csv_path: Path, event_name: str) -> int:
    """Remove rows for event_name from csv_path. Returns number of rows removed."""
    rows = bmi._read_csv_rows(csv_path)
    before = len(rows)
    kept = [r for r in rows if str(r.get("Event Name", "")).strip() != event_name]
    removed = before - len(kept)
    if removed:
        bmi._write_csv_rows(csv_path, kept)
    return removed


def _merge_into_manual_postprocessed(bmi, new_rows: List[Dict[str, str]]) -> Tuple[int, int]:
    """Replace-or-append by Event Name into database/data/tournament_manual_postprocessed.csv."""
    MANUAL_TOURNAMENT_CSV.parent.mkdir(parents=True, exist_ok=True)
    existing_rows = bmi._read_csv_rows(MANUAL_TOURNAMENT_CSV)
    target_event_names = {row["Event Name"] for row in new_rows}
    kept_rows = [row for row in existing_rows if row.get("Event Name", "") not in target_event_names]
    merged_rows = kept_rows + new_rows

    merged_rows.sort(
        key=lambda r: (
            r.get("Season", ""),
            r.get("Date", ""),
            r.get("Event Name", ""),
            int(r.get("Round Number", "0") or "0"),
            int(r.get("Game Number", "0") or "0"),
            r.get("Player", ""),
            r.get("Player ID", ""),
        )
    )
    bmi._write_csv_rows(MANUAL_TOURNAMENT_CSV, merged_rows)
    return len(existing_rows), len(merged_rows)


def _load_bayerische_import_module():
    """Load sibling import module (no package __init__)."""
    path = ROOT / "database" / "input" / "import_bayerische_meisterschaft_xlsx.py"
    name = "import_bayerische_meisterschaft_xlsx"
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).casefold()


def _name_variants(raw: str) -> List[str]:
    raw = (raw or "").strip()
    if not raw:
        return []
    variants = {_norm_ws(raw)}
    if "," in raw:
        parts = [p.strip() for p in raw.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            variants.add(_norm_ws(f"{parts[1]} {parts[0]}"))
            variants.add(_norm_ws(f"{parts[0]}, {parts[1]}"))
    else:
        tokens = raw.split()
        if len(tokens) >= 2:
            first_rest = " ".join(tokens[:-1])
            last = tokens[-1]
            variants.add(_norm_ws(f"{last}, {first_rest}"))
            # Surname-first token (e.g. "Luu Vinh Duc" <-> "Luu, Vinh Duc")
            variants.add(_norm_ws(f"{tokens[0]}, {' '.join(tokens[1:])}"))
    return list(variants)


def _as_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    t = str(value).strip().replace(",", ".")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _as_int_score(value: object) -> Optional[int]:
    f = _as_float(value)
    if f is None:
        return None
    if math.isnan(f):
        return None
    return int(round(f))


def _arith_round_int_pins(x: float) -> int:
    """Nearest pins integer, half up (non‑negative domain for handicap)."""
    if not math.isfinite(x) or x <= 0:
        return 0
    return math.floor(x + 0.5)


def _handicap_per_game_pins(reference: float, ap_avg: float) -> int:
    diff = reference - ap_avg
    if diff <= 0:
        return 0
    return _arith_round_int_pins(0.7 * diff)


def _build_player_id_lookup(league_csv: Path) -> Dict[str, str]:
    """
    Returns best Player ID per normalized name across league rows.
    Chooses most frequent Player ID per normalized name.
    """
    if not league_csv.is_file():
        return {}

    id_counts: Dict[str, Counter] = {}
    with league_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            if str(row.get("Input Data", "")).strip().lower() not in ("true", "1", "yes"):
                continue
            name = str(row.get("Player", "") or "").strip()
            if not name or name.lower() == "team total":
                continue
            pid = str(row.get("Player ID", "") or "").strip()
            if not pid or pid == "0":
                continue
            for key in _name_variants(name):
                id_counts.setdefault(key, Counter())[pid] += 1

    best: Dict[str, str] = {}
    for key, ctr in id_counts.items():
        if ctr:
            best[key] = ctr.most_common(1)[0][0]
    return best


def _resolve_player_id(
    display_name: str,
    lookup: Dict[str, str],
) -> Tuple[str, Optional[str]]:
    for key in _name_variants(display_name):
        pid = lookup.get(key)
        if pid:
            return pid, key
    return "", None


def _season_label_from_calendar_year(year: int) -> str:
    prev_yy = (year - 1) % 100
    curr_yy = year % 100
    return f"{prev_yy:02d}/{curr_yy:02d}"


def _iter_data_rows(ws, first_data_row: int):
    row_idx = first_data_row
    max_row = ws.max_row or first_data_row
    while row_idx <= max_row:
        name_cell = ws.cell(row=row_idx, column=COL_NAME).value
        name = str(name_cell).strip() if name_cell is not None else ""
        if not name or name.lower() in ("teilnehmer", "name", "player"):
            row_idx += 1
            continue
        yield row_idx, name
        row_idx += 1


def _extract_rows_for_sheet(
    ws,
    *,
    season: str,
    event_date: str,
    location: str,
    player_lookup: Dict[str, str],
    first_data_row: int,
) -> Tuple[List[Dict[str, str]], List[str]]:
    clean_rows: List[Dict[str, str]] = []
    unmatched: List[str] = []

    for row_idx, display_name in _iter_data_rows(ws, first_data_row):
        ref_f = _as_float(ws.cell(row=row_idx, column=COL_REF_HANDICAP).value)
        ap_f = _as_float(ws.cell(row=row_idx, column=COL_AP_AVG).value)
        if ref_f is not None and ap_f is not None:
            h_per_game = _handicap_per_game_pins(ref_f, ap_f)
        else:
            h_per_game = 0
        pid, _matched_key = _resolve_player_id(display_name, player_lookup)
        if not pid:
            unmatched.append(display_name)
            pid = f"UNK{hashlib.md5(display_name.encode('utf-8')).hexdigest()[:10]}"

        for set_idx in range(NUM_SET_BLOCKS):
            base_col = COL_FIRST_SET + set_idx * SET_BLOCK_WIDTH
            scratch = [
                _as_int_score(ws.cell(row=row_idx, column=base_col + j).value) for j in range(GAMES_PER_SET)
            ]
            if all(v is None for v in scratch):
                continue
            for game_idx, score in enumerate(scratch):
                if score is None:
                    continue
                row_out: Dict[str, str] = {
                    "Season": season,
                    "Date": event_date,
                    "Location": location,
                    "Event Type": "tournament",
                    "Event Name": EVENT_NAME,
                    "Round Number": str(set_idx + 1),
                    "Round Name": f"Set {set_idx + 1}",
                    "Player": display_name,
                    "Player ID": pid,
                    "Club": "",
                    "Game Number": str(game_idx),
                    "Score": str(score),
                    "Handicap": str(h_per_game),
                }
                if ref_f is not None:
                    row_out["Handicap Reference"] = str(int(round(ref_f)))
                if ap_f is not None:
                    ap_s = f"{float(ap_f):.4f}".rstrip("0").rstrip(".")
                    row_out["A Priori Average"] = ap_s if ap_s else str(float(ap_f))
                clean_rows.append(row_out)

    return clean_rows, unmatched


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Import Clubmeisterschaft Donaubowler XLSX.")
    p.add_argument(
        "--xlsx",
        type=str,
        default="",
        help="Path to a single .xlsx file (default: first *.xlsx in --input-dir).",
    )
    p.add_argument(
        "--input-dir",
        type=str,
        default=str(DEFAULT_INPUT_DIR),
        help="Directory containing Clubmeisterschaft workbook(s).",
    )
    p.add_argument(
        "--output",
        type=str,
        default=str(DEFAULT_OUTPUT),
        help="Batch postprocessed CSV output path.",
    )
    p.add_argument(
        "--league-csv",
        type=str,
        default=str(DEFAULT_LEAGUE_CSV),
        help="Merged league CSV for Player ID lookup.",
    )
    p.add_argument("--season", type=str, default="", help="Season label, e.g. 25/26 (default: from --year).")
    p.add_argument("--year", type=int, default=2026, help="Calendar year for default season label.")
    p.add_argument("--date", type=str, default="2026-05-15", help="ISO date for tournament rows.")
    p.add_argument("--location", type=str, default="", help="Venue / location string.")
    p.add_argument("--sheet", type=str, default="", help="Worksheet name (default: first sheet).")
    p.add_argument("--first-data-row", type=int, default=2, help="First Excel row index containing player data.")
    return p


def main() -> None:
    args = build_parser().parse_args()
    input_dir = Path(args.input_dir).resolve()
    league_csv = Path(args.league_csv).resolve()

    if args.xlsx:
        xlsx_path = Path(args.xlsx).resolve()
    else:
        candidates = sorted(
            p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~") and not p.name.startswith(".~")
        )
        if not candidates:
            raise FileNotFoundError(f"No .xlsx files in {input_dir}")
        xlsx_path = candidates[0]

    if not xlsx_path.is_file():
        raise FileNotFoundError(xlsx_path)

    season = (args.season or "").strip() or _season_label_from_calendar_year(int(args.year))

    bmi = _load_bayerische_import_module()
    player_lookup = _build_player_id_lookup(league_csv)

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = (args.sheet or "").strip() or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not in workbook. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    clean_rows, unmatched = _extract_rows_for_sheet(
        ws,
        season=season,
        event_date=str(args.date).strip(),
        location=str(args.location or "").strip(),
        player_lookup=player_lookup,
        first_data_row=int(args.first_data_row),
    )

    if not clean_rows:
        raise ValueError(f"No score rows extracted from {xlsx_path}")

    postprocessed = bmi._postprocess(clean_rows)
    output_path = Path(args.output).resolve()
    bmi._write_csv_rows(output_path, postprocessed)

    removed_from_gf = 0
    if GF_REGIONAL_TOURNAMENT_CSV.is_file():
        removed_from_gf = _strip_event_from_path(bmi, GF_REGIONAL_TOURNAMENT_CSV, EVENT_NAME)
    before_manual, after_manual = _merge_into_manual_postprocessed(bmi, postprocessed)
    _rebuild_player_hybrid_local()

    print(f"Source: {xlsx_path} (sheet {sheet_name!r})")
    print(f"Season: {season}  Event: {EVENT_NAME}")
    print(f"Wrote: {output_path} ({len(postprocessed)} postprocessed rows)")
    if removed_from_gf:
        print(
            f"Removed {removed_from_gf} rows for this event from GF export "
            f"({GF_REGIONAL_TOURNAMENT_CSV.name})"
        )
    print(
        f"Manual tournament CSV: {MANUAL_TOURNAMENT_CSV} "
        f"({before_manual} -> {after_manual} rows)"
    )
    print("Rebuilt player hybrid: database/data/player_stats_merged_plus_tournaments.csv")
    uniq_unmatched = sorted({n for n in unmatched if n})
    if uniq_unmatched:
        print(f"Warning: {len(uniq_unmatched)} player(s) without league ID match:")
        for n in uniq_unmatched[:30]:
            print(f"  - {n}")
        if len(uniq_unmatched) > 30:
            print(f"  ... and {len(uniq_unmatched) - 30} more")


if __name__ == "__main__":
    main()
