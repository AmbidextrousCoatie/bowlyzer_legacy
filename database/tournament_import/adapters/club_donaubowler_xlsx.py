"""Clubmeisterschaft Donaubowler XLSX adapter."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from typing import Dict, List

from database.paths import REPO_ROOT
from database.tournament_import.config import ImportEntry
from database.tournament_import.schema import season_label_from_calendar_year
from openpyxl import load_workbook


def _load_club_module():
    path = REPO_ROOT / "database" / "input" / "import_clubmeisterschaft_donaubowler_xlsx.py"
    name = "import_clubmeisterschaft_donaubowler_xlsx"
    mod = sys.modules.get(name)
    if mod is not None:
        # Reload so adapter picks up importer edits in long-lived processes.
        import importlib

        return importlib.reload(mod)
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


class ClubDonaubowlerXlsxAdapter:
    format_id = "club_donaubowler_xlsx"

    def parse(self, source: Path, entry: ImportEntry) -> List[Dict[str, str]]:
        club = _load_club_module()
        opts = entry.options

        if source.is_dir():
            xlsx_path = club._pick_qualifying_xlsx(source)
            finale_path = None if opts.get("no_finale") else club._pick_finale_xlsx(source)
        else:
            xlsx_path = source
            finale_path = None
            if not opts.get("no_finale"):
                finale_path = club._pick_finale_xlsx(source.parent)

        if not xlsx_path.is_file():
            raise FileNotFoundError(xlsx_path)

        year = int(opts.get("year", 2026))
        season = str(opts.get("season") or "").strip() or season_label_from_calendar_year(year)
        event_date = str(opts.get("date") or "2026-05-15").strip()
        location = str(opts.get("location") or "").strip()
        sheet = str(opts.get("sheet") or "").strip()
        first_data_row = int(opts.get("first_data_row", 2))
        league_csv = Path(str(opts.get("league_csv") or club.DEFAULT_LEAGUE_CSV)).resolve()

        player_lookup = club._build_player_id_lookup(league_csv)
        wb = load_workbook(xlsx_path, data_only=True)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not in workbook. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        clean_rows, _unmatched = club._extract_rows_for_sheet(
            ws,
            season=season,
            event_date=event_date,
            location=location,
            player_lookup=player_lookup,
            first_data_row=first_data_row,
        )
        preferred_names_by_id: Dict[str, str] = {}
        for row in clean_rows:
            pid = str(row.get("Player ID") or "").strip()
            pname = str(row.get("Player") or "").strip()
            if pid and pname and pid not in preferred_names_by_id:
                preferred_names_by_id[pid] = pname

        if finale_path and finale_path.is_file():
            fwb = load_workbook(finale_path, data_only=True)
            fws = fwb[fwb.sheetnames[0]]
            finale_rows, _ = club._extract_finale_rows_for_sheet(
                fws,
                season=season,
                event_date=event_date,
                location=location,
                player_lookup=player_lookup,
                preferred_names_by_id=preferred_names_by_id,
            )
            clean_rows = clean_rows + finale_rows

        if not clean_rows:
            raise ValueError(f"No score rows extracted from {xlsx_path}")
        return clean_rows
