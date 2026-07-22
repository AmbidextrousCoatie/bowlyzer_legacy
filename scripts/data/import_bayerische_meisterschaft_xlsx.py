"""
Import Bayerische Meisterschaft XLSX files into tournament postprocessed CSV.

Expected workbook structure:
- Sheet "Optionen": B3 (name), B4 (subgroup), B5 (season year), B7 (date range)
- Sheet "Vorrunde": rows from 6, columns A..H (name, id, six games), I optional total
- Sheet "Zwischenlauf": rows from 6, columns A..H (name, id, six games), I optional total

Output:
- Per-import postprocessed CSV under database/work/tournaments/staging/
- Merged append/update into database/work/gf/gf_tournaments_2026__combined_postprocessed.csv
- Rebuilds player hybrid via publish pipeline
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database.tournament_import.io import read_csv_rows as _read_csv_rows
from database.tournament_import.io import write_csv_rows as _write_csv_rows
from database.tournament_import.postprocess import postprocess_rows as _postprocess
from database.tournament_import.schema import POSTPROCESSED_HEADERS

from database.paths import get_data_dir, gf_tournaments_combined_postprocessed_csv, tournament_staging_dir

GF_COMBINED_POSTPROCESSED = gf_tournaments_combined_postprocessed_csv()
DEFAULT_OUTPUT = tournament_staging_dir() / "tournament_bayerische_meisterschaft_2026_postprocessed.csv"
KO_CONFIG_PATH = get_data_dir() / "tournament_ko_config.json"


@dataclass(frozen=True)
class TournamentMeta:
    season: str
    event_name: str
    location: str
    dates: Dict[int, str]


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_int(value: object, default: int = 0) -> int:
    try:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return default
        return int(float(value))
    except Exception:
        return default


def _as_float(value: object) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _as_int_score(value: object) -> Optional[int]:
    """Scratch pins from a game cell; None when the cell is blank (not bowled)."""
    parsed = _as_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def _season_label(year: int) -> str:
    # Keep same style used by SBM/NBM GF exports, e.g. 2026 -> 25/26.
    prev_yy = (year - 1) % 100
    curr_yy = year % 100
    return f"{prev_yy:02d}/{curr_yy:02d}"


def _parse_date_token(token: str, season_year: int) -> str:
    token = token.strip()
    if "." not in token:
        raise ValueError(f"Invalid date token '{token}', expected d.m")
    day_s, month_s = token.split(".", 1)
    day = int(day_s)
    month = int(month_s)
    return date(season_year, month, day).isoformat()


def _parse_date_range(raw: str, season_year: int) -> Dict[int, str]:
    value = raw.replace("bis", "-")
    parts = [p.strip() for p in value.split("-") if p.strip()]
    if len(parts) < 2:
        raise ValueError(f"Invalid date range '{raw}'")
    return {1: _parse_date_token(parts[0], season_year), 2: _parse_date_token(parts[1], season_year)}


def _read_optionen_meta(workbook_path: Path) -> TournamentMeta:
    wb = load_workbook(workbook_path, data_only=True)
    if "Optionen" not in wb.sheetnames:
        raise ValueError(f"Workbook missing 'Optionen' sheet: {workbook_path}")

    ws = wb["Optionen"]
    event_base = _clean_text(ws["B3"].value)
    subgroup = _clean_text(ws["B4"].value)
    season_year = _as_int(ws["B5"].value)
    date_range = _clean_text(ws["B7"].value)
    location = _clean_text(ws["B12"].value)

    if not event_base or not subgroup or season_year <= 0 or not date_range:
        raise ValueError(f"Incomplete Optionen metadata in '{workbook_path.name}'")

    event_name = f"{event_base} - {subgroup}"
    season = _season_label(season_year)
    dates = _parse_date_range(date_range, season_year)
    return TournamentMeta(season=season, event_name=event_name, location=location, dates=dates)


def _iter_round_rows(workbook_path: Path, sheet_name: str) -> Iterable[Tuple[str, str, List[Optional[int]]]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows: List[Tuple[str, str, List[Optional[int]]]] = []
    row_idx = 6
    consecutive_empty = 0
    # Some sheets contain visual spacer rows in the middle; only stop after
    # a longer empty run near the real end of data.
    max_consecutive_empty = 20
    max_scan_row = 800
    while row_idx <= max_scan_row:
        name = _clean_text(ws.cell(row=row_idx, column=1).value)
        pid = _as_int(ws.cell(row=row_idx, column=2).value)
        if not name and pid <= 0:
            consecutive_empty += 1
            if consecutive_empty >= max_consecutive_empty:
                break
            row_idx += 1
            continue
        consecutive_empty = 0
        if name and pid > 0:
            games = [_as_int_score(ws.cell(row=row_idx, column=col).value) for col in range(3, 9)]
            rows.append((name, str(pid), games))
        row_idx += 1
    return rows


def _build_round_rows(meta: TournamentMeta, workbook_path: Path) -> Tuple[List[Dict[str, str]], Dict[str, str]]:
    round_map = {
        1: ("Vorrunde", "Vorlauf"),
        2: ("Zwischenlauf", "Zwischenlauf"),
    }
    rows: List[Dict[str, str]] = []
    player_id_by_name: Dict[str, str] = {}
    for round_number, (sheet_name, round_name) in round_map.items():
        source_rows = list(_iter_round_rows(workbook_path, sheet_name))
        for player_name, player_id, game_scores in source_rows:
            player_id_by_name[player_name.strip().casefold()] = str(player_id)
            for game_idx, score in enumerate(game_scores):
                if score is None:
                    continue
                rows.append(
                    {
                        "Season": meta.season,
                        "Date": meta.dates[round_number],
                        "Location": meta.location,
                        "Event Type": "tournament",
                        "Event Name": meta.event_name,
                        "Round Number": str(round_number),
                        "Round Name": round_name,
                        "Player": player_name,
                        "Player ID": player_id,
                        "Club": "",
                        "Game Number": str(game_idx),
                        "Score": str(score),
                        "Handicap": "0",
                    }
                )
    return rows, player_id_by_name


def _is_number(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    try:
        float(text)
        return True
    except Exception:
        return False


def _load_ko_finale_series(season: str, event_name: str) -> str:
    """Read ko_finale_series from tournament_ko_config.json (defaults to bo3_pins)."""
    if not KO_CONFIG_PATH.is_file():
        return "bo3_pins"
    try:
        data = json.loads(KO_CONFIG_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return "bo3_pins"
    block = data.get(f"{season}||{event_name}") or {}
    return str(block.get("ko_finale_series") or "bo3_pins")


def _parse_ko_games(
    ws,
    start_row: int,
    active_section: str = "",
    ko_finale_series: str = "bo3_pins",
) -> Tuple[List[Tuple[int, int]], int]:
    """
    Parse KO pin rows beginning at start_row.
    Returns (list of (left,right) played games, next_row_after_match_block).

    bo3_pins: up to three single-game rows per match; skip impossible singles (>300 either side).
    scratch_total_2g: at most two single-game rows; skip two-game scratch totals (often <300 per side,
    e.g. 295) by matching running sums within the match block.
    In *finale*, skip tiny pairs like 2:1 that are "games won" summaries, not pinfall.
    """
    games: List[Tuple[int, int]] = []
    row = start_row
    inspected = 0
    max_rows = 8  # game rows + optional total + optional "gewonne Spiele"
    sec = (active_section or "").strip().lower()
    while row <= ws.max_row and inspected < max_rows:
        inspected += 1
        left_raw = ws.cell(row=row, column=3).value
        colon = _clean_text(ws.cell(row=row, column=4).value)
        right_raw = ws.cell(row=row, column=5).value
        b_val = _clean_text(ws.cell(row=row, column=2).value).lower()

        if "gewonne spiele" in b_val:
            row += 1
            break
        if colon != ":":
            break

        left_is_num = _is_number(left_raw)
        right_is_num = _is_number(right_raw)
        if left_is_num and right_is_num:
            left_num = int(float(left_raw))
            right_num = int(float(right_raw))
            if left_num == 0 and right_num == 0:
                row += 1
                continue
            # Single-game scratch cap (either side); catches match totals for bo3 as well.
            if left_num > 300 or right_num > 300:
                row += 1
                continue
            if ko_finale_series == "scratch_total_2g":
                if len(games) >= 2:
                    row += 1
                    continue
                if games:
                    sum_left = sum(g[0] for g in games)
                    sum_right = sum(g[1] for g in games)
                    if left_num == sum_left or right_num == sum_right:
                        row += 1
                        continue
            # Finale sometimes lists leg wins (e.g. 2:1) as a separate row — not pinfall.
            if sec == "finale" and left_num <= 10 and right_num <= 10:
                row += 1
                continue
            games.append((left_num, right_num))
            row += 1
            continue

        # Rows like "- : -" indicate no game played in this slot.
        left_text = _clean_text(left_raw)
        right_text = _clean_text(right_raw)
        if left_text == "-" and right_text == "-":
            row += 1
            continue
        break

    return games, row


def _is_walkover_opponent(name: str) -> bool:
    n = name.strip().lower()
    if not n or n == "-":
        return True
    if "nicht angetreten" in n:
        return True
    return False


def _no_show_display_name(raw_absent_cell: str) -> str:
    """
    Build CSV / UI label: real name when the sheet embeds it next to 'nicht angetreten',
    otherwise 'Nicht angetreten (No show)'. Always ends with ' (No show)' for clarity.
    """
    t = (raw_absent_cell or "").strip()
    tl = t.lower()
    if tl.endswith("(no show)") or re.search(r"\(\s*no\s*show\s*\)\s*$", tl):
        return t
    if "nicht angetreten" in tl:
        stripped = re.sub(r"[\s\-–—]*nicht\s+angetreten[\s\-–—]*", "", t, flags=re.I).strip(" -–—")
        if len(stripped) >= 2:
            base = stripped.strip()
            if not base.lower().endswith("(no show)"):
                return f"{base} (No show)"
            return base
        return "Nicht angetreten (No show)"
    if not t:
        return "Nicht angetreten (No show)"
    return f"{t} (No show)"


def _extract_ko_rows(
    meta: TournamentMeta,
    workbook_path: Path,
    player_id_by_name: Dict[str, str],
) -> List[Dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    if "KO-Finale" not in wb.sheetnames:
        return []
    ws = wb["KO-Finale"]

    round_number = 3
    round_name = "KO-Finale"
    round_date = meta.dates.get(2, "")
    ko_finale_series = _load_ko_finale_series(meta.season, meta.event_name)

    section_labels = {"viertelfinale", "halbfinale", "finale"}
    ko_rows: List[Dict[str, str]] = []
    game_number = 0
    row = 1
    while row <= ws.max_row:
        section_cell = _clean_text(ws.cell(row=row, column=3).value).lower()
        if section_cell not in section_labels:
            row += 1
            continue
        active_section = section_cell
        row += 1
        # Parse section until next section anchor.
        while row <= ws.max_row:
            next_section = _clean_text(ws.cell(row=row, column=3).value).lower()
            if next_section in section_labels:
                break
            left_name = _clean_text(ws.cell(row=row, column=2).value)
            right_name = _clean_text(ws.cell(row=row, column=6).value)
            winner_marker = _clean_text(ws.cell(row=row, column=9).value)
            left_lower = left_name.lower()
            # Halbfinale rows sometimes omit the opponent cell when it is a no-show; treat like walkover.
            allow_missing_right = active_section == "halbfinale" and left_name and not right_name
            if (
                left_name
                and (right_name or allow_missing_right)
                and ":" == _clean_text(ws.cell(row=row, column=4).value)
                and "gewonne spiele" not in left_lower
                and not left_lower.startswith("platz ")
                and "vorrunde platz" not in left_lower
                and "sieger" not in left_lower
            ):
                games, next_row = _parse_ko_games(ws, row, active_section, ko_finale_series)
                left_id = player_id_by_name.get(left_name.strip().casefold(), "")
                right_id = player_id_by_name.get(right_name.strip().casefold(), "")
                if _is_walkover_opponent(right_name):
                    absent_display = _no_show_display_name(right_name)
                    # No-show: emit a single 0–0 game pair tagged KO_WO so bracket UI can show walkover
                    # without counting pins in downstream stats (scores are 0).
                    ko_rows.append(
                        {
                            "Season": meta.season,
                            "Date": round_date,
                            "Location": meta.location,
                            "Event Type": "tournament",
                            "Event Name": meta.event_name,
                            "Round Number": str(round_number),
                            "Round Name": round_name,
                            "Player": left_name,
                            "Player ID": left_id,
                            "Club": "KO_WO",
                            "Game Number": str(game_number),
                            "Score": "0",
                            "Handicap": "0",
                        }
                    )
                    ko_rows.append(
                        {
                            "Season": meta.season,
                            "Date": round_date,
                            "Location": meta.location,
                            "Event Type": "tournament",
                            "Event Name": meta.event_name,
                            "Round Number": str(round_number),
                            "Round Name": round_name,
                            "Player": absent_display,
                            "Player ID": "",
                            "Club": "KO_WO",
                            "Game Number": str(game_number),
                            "Score": "0",
                            "Handicap": "0",
                        }
                    )
                    game_number += 1
                elif _is_walkover_opponent(left_name):
                    absent_display = _no_show_display_name(left_name)
                    ko_rows.append(
                        {
                            "Season": meta.season,
                            "Date": round_date,
                            "Location": meta.location,
                            "Event Type": "tournament",
                            "Event Name": meta.event_name,
                            "Round Number": str(round_number),
                            "Round Name": round_name,
                            "Player": absent_display,
                            "Player ID": "",
                            "Club": "KO_WO",
                            "Game Number": str(game_number),
                            "Score": "0",
                            "Handicap": "0",
                        }
                    )
                    ko_rows.append(
                        {
                            "Season": meta.season,
                            "Date": round_date,
                            "Location": meta.location,
                            "Event Type": "tournament",
                            "Event Name": meta.event_name,
                            "Round Number": str(round_number),
                            "Round Name": round_name,
                            "Player": right_name,
                            "Player ID": right_id,
                            "Club": "KO_WO",
                            "Game Number": str(game_number),
                            "Score": "0",
                            "Handicap": "0",
                        }
                    )
                    game_number += 1
                else:
                    for left_score, right_score in games:
                        if left_score <= 0 or right_score <= 0:
                            continue
                        ko_rows.append(
                            {
                                "Season": meta.season,
                                "Date": round_date,
                                "Location": meta.location,
                                "Event Type": "tournament",
                                "Event Name": meta.event_name,
                                "Round Number": str(round_number),
                                "Round Name": round_name,
                                "Player": left_name,
                                "Player ID": left_id,
                                "Club": "",
                                "Game Number": str(game_number),
                                "Score": str(left_score),
                                "Handicap": "0",
                            }
                        )
                        ko_rows.append(
                            {
                                "Season": meta.season,
                                "Date": round_date,
                                "Location": meta.location,
                                "Event Type": "tournament",
                                "Event Name": meta.event_name,
                                "Round Number": str(round_number),
                                "Round Name": round_name,
                                "Player": right_name,
                                "Player ID": right_id,
                                "Club": "",
                                "Game Number": str(game_number),
                                "Score": str(right_score),
                                "Handicap": "0",
                            }
                        )
                        game_number += 1
                # If no played game rows were detected, still advance to avoid loops.
                row = max(next_row, row + 1)
                continue
            # Header rows (seed labels / summary rows) are skipped.
            if winner_marker.lower().startswith("platz "):
                row += 1
                continue
            row += 1
    return ko_rows


def _merge_into_combined(new_rows: List[Dict[str, str]]) -> Tuple[int, int]:
    existing_rows = _read_csv_rows(GF_COMBINED_POSTPROCESSED)
    target_event_names = {row["Event Name"] for row in new_rows}
    kept_rows = [row for row in existing_rows if row.get("Event Name", "") not in target_event_names]
    merged_rows = kept_rows + new_rows

    merged_rows.sort(
        key=lambda r: (
            r.get("Season", ""),
            r.get("Date", ""),
            r.get("Event Name", ""),
            int(r.get("Round Number", "0") or "0"),
            int(r.get("Game Number", "0") or "0"),
            r.get("Player", ""),
            r.get("Player ID", ""),
        )
    )
    _write_csv_rows(GF_COMBINED_POSTPROCESSED, merged_rows)
    return len(existing_rows), len(merged_rows)


def _rebuild_player_hybrid() -> None:
    # Reuse existing app config merge logic to keep schema in sync.
    from app.config.database_config import _build_player_merged_hybrid_csv

    _build_player_merged_hybrid_csv()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import Bayerische Meisterschaft XLSX files.")
    parser.add_argument(
        "--xlsx-dir",
        required=True,
        help="Directory containing Bayerische Meisterschaft XLSX files.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Output postprocessed CSV for this import batch.",
    )
    parser.add_argument(
        "--include-ko-finale",
        action="store_true",
        help="Include parsed KO-Finale best-of-3 bracket as round 3.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    xlsx_dir = Path(args.xlsx_dir).resolve()
    output_path = Path(args.output).resolve()
    if not xlsx_dir.is_dir():
        raise FileNotFoundError(f"XLSX directory not found: {xlsx_dir}")

    all_clean_rows: List[Dict[str, str]] = []
    for workbook_path in sorted(xlsx_dir.glob("*.xlsx")):
        meta = _read_optionen_meta(workbook_path)
        workbook_rows, player_id_by_name = _build_round_rows(meta, workbook_path)
        if args.include_ko_finale:
            workbook_rows.extend(_extract_ko_rows(meta, workbook_path, player_id_by_name))
        all_clean_rows.extend(workbook_rows)

    if not all_clean_rows:
        raise ValueError(f"No tournament rows found in '{xlsx_dir}'")

    postprocessed_rows = _postprocess(all_clean_rows)
    _write_csv_rows(output_path, postprocessed_rows)
    before_count, after_count = _merge_into_combined(postprocessed_rows)
    _rebuild_player_hybrid()

    event_names = sorted({row["Event Name"] for row in postprocessed_rows})
    print(f"Imported events: {event_names}")
    print(f"Wrote batch output: {output_path} ({len(postprocessed_rows)} rows)")
    print(
        "Updated combined tournaments CSV: "
        f"{GF_COMBINED_POSTPROCESSED} ({before_count} -> {after_count} rows)"
    )
    print("Rebuilt player hybrid CSV: database/data/player_stats_merged_plus_tournaments.csv")


if __name__ == "__main__":
    main()
