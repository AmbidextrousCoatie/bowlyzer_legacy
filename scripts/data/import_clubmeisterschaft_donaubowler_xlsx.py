"""
Import Clubmeisterschaft Donaubowler XLSX into tournament postprocessed CSV.

(Clubpokal is a separate team KO competition over multiple months — not this importer.)

Workbook layout (Clubmeisterschaft 2026 export):
- Col A: rank (ignored for import)
- Col B: player name
- Cols C onward: repeating blocks of 6 columns per set — 4 scratch games, handicap (sheet), total (sheet).
  We only emit scratch scores; handicap column uses computed per-game handicap (see below).
- Col AO (41): reference score for handicap
- Col AP (42): a priori average

Handicap per game: max(0, 0.7 * (reference_score - a_priori_average))  [pins, rounded half up to nearest int]

Player IDs: resolved from merged league data (CSV or Parquet) and ``players_registry.parquet``.

Output:
- Staging CSV under database/work/tournaments/staging/
- Merges into database/work/tournaments/tournament_manual_postprocessed.csv
- Removes this event from GF regional export if previously written there
- Rebuilds player hybrid via publish pipeline
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import math
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database.paths import (
    gf_tournaments_combined_postprocessed_csv,
    manual_tournament_postprocessed_csv,
    raw_input_dir,
    tournament_staging_dir,
)

DEFAULT_INPUT_DIR = raw_input_dir() / "clubmeisterschaft_donaubowler"
DEFAULT_OUTPUT = tournament_staging_dir() / "tournament_clubmeisterschaft_donaubowler_2026_postprocessed.csv"
DEFAULT_LEAGUE_CSV = ROOT / "database" / "data" / "league_results_merged.csv"
GF_REGIONAL_TOURNAMENT_CSV = gf_tournaments_combined_postprocessed_csv()
MANUAL_TOURNAMENT_CSV = manual_tournament_postprocessed_csv()

COL_NAME = 2
COL_FIRST_SET = 3
COL_REF_HANDICAP = 41  # AO
COL_AP_AVG = 42  # AP
SET_BLOCK_WIDTH = 6
GAMES_PER_SET = 4
NUM_SET_BLOCKS = 6  # six sets in stage 1

EVENT_NAME = "Clubmeisterschaft Donaubowler 2026"

# Finals workbook: "Last, First" / typos → preferred display (qualifying uses "First Last").
FINALE_NAME_ALIASES = {
    "harteil, volkmar": "Volkmar Hartfeil",
    "hartfeil, volkmar": "Volkmar Hartfeil",
    "feller, christian": "Christian Feller",
    "obermeier, kurt": "Kurt Obermeier",
    "schneider, tobias": "Tobias Schneider",
    "rettinger, fabian": "Fabian Rettinger",
    "luu, vinh duc": "Luu Vinh Duc",
}

# Finals sheet columns (1-based): seed, name, hdc/game, then scratch games, then inkl-hdc mirrors.
FINALE_COL_SEED = 1
FINALE_COL_NAME = 2
FINALE_COL_HDC = 3
FINALE_SCRATCH_COLS = {
    "elim_1": 4,
    "elim_2": 5,
    "sl1": 6,
    "sl2": 7,
    "f1": 8,
    "f2": 9,
    # Optional 3rd final game if present later
    "f3": None,
}
FINALE_ROUND_ELIM = 7
FINALE_ROUND_STEPLADDER = 8
FINALE_ROUND_FINAL = 9

# Workbook-only fields for change detection (no Player ID — league lookup must not affect fingerprint).
FINGERPRINT_VERSION = "clubmeisterschaft-sheet-v2"
FINGERPRINT_FIELDS = (
    "Season",
    "Date",
    "Location",
    "Event Name",
    "Round Number",
    "Round Name",
    "Player",
    "Game Number",
    "Score",
    "Handicap",
    "Handicap Reference",
    "A Priori Average",
)


def _rebuild_player_hybrid_local() -> None:
    """League + GF regional tournament export + manual tournament imports (matches database_config)."""
    merged_league = ROOT / "database" / "data" / "league_results_merged.csv"
    out_path = ROOT / "database" / "data" / "player_stats_merged_plus_tournaments.csv"

    league_rows: List[Dict[str, str]] = []
    if merged_league.is_file():
        with merged_league.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            league_rows = [{str(k): str(v or "") for k, v in row.items()} for row in reader]

    tournament_rows: List[Dict[str, str]] = []
    if GF_REGIONAL_TOURNAMENT_CSV.is_file():
        with GF_REGIONAL_TOURNAMENT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            tournament_rows.extend([{str(k): str(v or "") for k, v in row.items()} for row in reader])
    if MANUAL_TOURNAMENT_CSV.is_file():
        with MANUAL_TOURNAMENT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            tournament_rows.extend([{str(k): str(v or "") for k, v in row.items()} for row in reader])

    headers = sorted({k for r in (league_rows + tournament_rows) for k in r.keys()})
    if not headers:
        from database.conversion.bowlingbayern_legacy_core import OUTPUT_HEADERS

        headers = list(OUTPUT_HEADERS)

    out_rows: List[Dict[str, str]] = []
    for row in league_rows:
        merged = {h: str(row.get(h, "")) for h in headers}
        if not str(merged.get("Event Type", "")).strip():
            merged["Event Type"] = "league"
        if not str(merged.get("Event Name", "")).strip():
            merged["Event Name"] = str(merged.get("League", "")).strip()
        if not str(merged.get("Club", "")).strip():
            merged["Club"] = str(merged.get("Team", "")).strip()
        out_rows.append(merged)

    for row in tournament_rows:
        merged = {h: str(row.get(h, "")) for h in headers}
        merged["Input Data"] = "True"
        merged["Computed Data"] = "False"
        out_rows.append(merged)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(out_rows)


def _strip_event_from_path(bmi, csv_path: Path, event_name: str) -> int:
    """Remove rows for event_name from csv_path. Returns number of rows removed."""
    rows = bmi._read_csv_rows(csv_path)
    before = len(rows)
    kept = [r for r in rows if str(r.get("Event Name", "")).strip() != event_name]
    removed = before - len(kept)
    if removed:
        bmi._write_csv_rows(csv_path, kept)
    return removed


def _merge_into_manual_postprocessed(bmi, new_rows: List[Dict[str, str]]) -> Tuple[int, int]:
    """Replace-or-append by Event Name into database/data/tournament_manual_postprocessed.csv."""
    MANUAL_TOURNAMENT_CSV.parent.mkdir(parents=True, exist_ok=True)
    existing_rows = bmi._read_csv_rows(MANUAL_TOURNAMENT_CSV)
    target_event_names = {row["Event Name"] for row in new_rows}
    kept_rows = [row for row in existing_rows if row.get("Event Name", "") not in target_event_names]
    merged_rows = kept_rows + new_rows

    merged_rows.sort(
        key=lambda r: (
            r.get("Season", ""),
            r.get("Date", ""),
            r.get("Event Name", ""),
            int(r.get("Round Number", "0") or "0"),
            int(r.get("Game Number", "0") or "0"),
            r.get("Player", ""),
            r.get("Player ID", ""),
        )
    )
    bmi._write_csv_rows(MANUAL_TOURNAMENT_CSV, merged_rows)
    return len(existing_rows), len(merged_rows)


def _load_bayerische_import_module():
    """Load sibling import module (no package __init__)."""
    path = ROOT / "database" / "input" / "import_bayerische_meisterschaft_xlsx.py"
    name = "import_bayerische_meisterschaft_xlsx"
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).casefold()


def _name_variants(raw: str) -> List[str]:
    raw = (raw or "").strip()
    if not raw:
        return []
    variants = {_norm_ws(raw)}
    if "," in raw:
        parts = [p.strip() for p in raw.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            variants.add(_norm_ws(f"{parts[1]} {parts[0]}"))
            variants.add(_norm_ws(f"{parts[0]}, {parts[1]}"))
    else:
        tokens = raw.split()
        if len(tokens) >= 2:
            first_rest = " ".join(tokens[:-1])
            last = tokens[-1]
            variants.add(_norm_ws(f"{last}, {first_rest}"))
            # Surname-first token (e.g. "Luu Vinh Duc" <-> "Luu, Vinh Duc")
            variants.add(_norm_ws(f"{tokens[0]}, {' '.join(tokens[1:])}"))
    return list(variants)


def _as_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    t = str(value).strip().replace(",", ".")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _as_int_score(value: object) -> Optional[int]:
    f = _as_float(value)
    if f is None:
        return None
    if math.isnan(f):
        return None
    return int(round(f))


def _arith_round_int_pins(x: float) -> int:
    """Nearest pins integer, half up (non‑negative domain for handicap)."""
    if not math.isfinite(x) or x <= 0:
        return 0
    return math.floor(x + 0.5)


def _handicap_per_game_pins(reference: float, ap_avg: float) -> int:
    diff = reference - ap_avg
    if diff <= 0:
        return 0
    return _arith_round_int_pins(0.7 * diff)


def _lookup_from_league_rows(rows: Iterable[Mapping[str, str]]) -> Dict[str, str]:
    """Best Player ID per normalized name from league row dicts."""
    id_counts: Dict[str, Counter] = {}
    for row in rows:
        if str(row.get("Input Data", "")).strip().lower() not in ("true", "1", "yes"):
            continue
        name = str(row.get("Player", "") or "").strip()
        if not name or name.lower() == "team total":
            continue
        pid = str(row.get("Player ID", "") or "").strip()
        if not pid or pid == "0":
            continue
        for key in _name_variants(name):
            id_counts.setdefault(key, Counter())[pid] += 1

    best: Dict[str, str] = {}
    for key, ctr in id_counts.items():
        if ctr:
            best[key] = ctr.most_common(1)[0][0]
    return best


def _lookup_from_league_csv(league_csv: Path) -> Dict[str, str]:
    if not league_csv.is_file():
        return {}
    with league_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        return _lookup_from_league_rows(reader)


def _lookup_from_league_parquet(logical_csv: Path) -> Dict[str, str]:
    try:
        from data_access.parquet_sidecar import data_file_exists, resolve_load_path
    except ImportError:
        return {}
    if not data_file_exists(logical_csv):
        return {}
    load_path = resolve_load_path(logical_csv)
    if load_path.suffix.lower() != ".parquet":
        return {}
    import pandas as pd

    df = pd.read_parquet(load_path)
    rows = [{str(k): str(v if v is not None else "") for k, v in row.items()} for row in df.to_dict("records")]
    return _lookup_from_league_rows(rows)


def _lookup_from_players_registry() -> Dict[str, str]:
    try:
        from data_access.players_registry import (
            candidate_names_for_entry,
            load_players_registry_df,
            registry_lookup_by_id,
        )
    except ImportError:
        return {}
    registry = load_players_registry_df()
    if registry is None or registry.empty:
        return {}
    out: Dict[str, str] = {}
    for pid, entry in registry_lookup_by_id(registry).items():
        for label in candidate_names_for_entry(entry):
            for key in _name_variants(label):
                out.setdefault(key, pid)
    return out


def _build_player_id_lookup(league_csv: Path) -> Dict[str, str]:
    """
    Player ID lookup for club tournament import.

    VPS often has league Parquet only (no CSV). Falls back to ``players_registry.parquet``.
    """
    lookup = _lookup_from_league_csv(league_csv)
    if not lookup:
        lookup = _lookup_from_league_parquet(league_csv)
    registry_lookup = _lookup_from_players_registry()
    for key, pid in registry_lookup.items():
        lookup.setdefault(key, pid)
    return lookup


def _resolve_player_id(
    display_name: str,
    lookup: Dict[str, str],
) -> Tuple[str, Optional[str]]:
    for key in _name_variants(display_name):
        pid = lookup.get(key)
        if pid:
            return pid, key
    return "", None


def _season_label_from_calendar_year(year: int) -> str:
    prev_yy = (year - 1) % 100
    curr_yy = year % 100
    return f"{prev_yy:02d}/{curr_yy:02d}"


def _iter_data_rows(ws, first_data_row: int):
    row_idx = first_data_row
    max_row = ws.max_row or first_data_row
    while row_idx <= max_row:
        name_cell = ws.cell(row=row_idx, column=COL_NAME).value
        name = str(name_cell).strip() if name_cell is not None else ""
        if not name or name.lower() in ("teilnehmer", "name", "player"):
            row_idx += 1
            continue
        yield row_idx, name
        row_idx += 1


def _extract_rows_for_sheet(
    ws,
    *,
    season: str,
    event_date: str,
    location: str,
    player_lookup: Dict[str, str],
    first_data_row: int,
) -> Tuple[List[Dict[str, str]], List[str]]:
    clean_rows: List[Dict[str, str]] = []
    unmatched: List[str] = []

    for row_idx, display_name in _iter_data_rows(ws, first_data_row):
        ref_f = _as_float(ws.cell(row=row_idx, column=COL_REF_HANDICAP).value)
        ap_f = _as_float(ws.cell(row=row_idx, column=COL_AP_AVG).value)
        if ref_f is not None and ap_f is not None:
            h_per_game = _handicap_per_game_pins(ref_f, ap_f)
        else:
            h_per_game = 0
        pid, _matched_key = _resolve_player_id(display_name, player_lookup)
        if not pid:
            unmatched.append(display_name)
            pid = f"UNK{hashlib.md5(display_name.encode('utf-8')).hexdigest()[:10]}"

        for set_idx in range(NUM_SET_BLOCKS):
            base_col = COL_FIRST_SET + set_idx * SET_BLOCK_WIDTH
            scratch = [
                _as_int_score(ws.cell(row=row_idx, column=base_col + j).value) for j in range(GAMES_PER_SET)
            ]
            if all(v is None for v in scratch):
                continue
            for game_idx, score in enumerate(scratch):
                if score is None:
                    continue
                row_out: Dict[str, str] = {
                    "Season": season,
                    "Date": event_date,
                    "Location": location,
                    "Event Type": "tournament",
                    "Event Name": EVENT_NAME,
                    "Round Number": str(set_idx + 1),
                    "Round Name": f"Set {set_idx + 1}",
                    "Player": display_name,
                    "Player ID": pid,
                    "Club": "",
                    "Game Number": str(game_idx),
                    "Score": str(score),
                    "Handicap": str(h_per_game),
                }
                if ref_f is not None:
                    row_out["Handicap Reference"] = str(int(round(ref_f)))
                if ap_f is not None:
                    ap_s = f"{float(ap_f):.4f}".rstrip("0").rstrip(".")
                    row_out["A Priori Average"] = ap_s if ap_s else str(float(ap_f))
                clean_rows.append(row_out)

    return clean_rows, unmatched


def _canonicalize_finale_name(raw: str) -> str:
    name = (raw or "").strip()
    if not name:
        return name
    alias = FINALE_NAME_ALIASES.get(_norm_ws(name))
    if alias:
        return alias
    # "Last, First" → "First Last"
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            return f"{parts[1]} {parts[0]}"
    return name


def _extract_finale_rows_for_sheet(
    ws,
    *,
    season: str,
    event_date: str,
    location: str,
    player_lookup: Dict[str, str],
    preferred_names_by_id: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict[str, str]], List[str]]:
    """
    Parse Clubpokal DB 2026 Finale.xlsx layout:
    row1 banners (HDC / scratch / inkl hdc), row2 headers, data from row3.
    Stores scratch in Score and per-game HDC in Handicap (decisions = Score+Handicap).
    """
    preferred_names_by_id = preferred_names_by_id or {}
    clean_rows: List[Dict[str, str]] = []
    unmatched: List[str] = []

    # Optional Finals 3 column if header present in row 2.
    f3_col: Optional[int] = None
    for c in range(1, (ws.max_column or 1) + 1):
        hdr = str(ws.cell(2, c).value or "").strip().casefold()
        if hdr in ("finals 3", "finale 3", "final 3"):
            f3_col = c
            break

    game_specs: List[Tuple[str, str, int, int]] = [
        # key, round_name, round_number, game_number
        ("elim_1", "KO Eliminierung", FINALE_ROUND_ELIM, 0),
        ("elim_2", "KO Eliminierung", FINALE_ROUND_ELIM, 1),
        ("sl1", "KO Stepladder", FINALE_ROUND_STEPLADDER, 2),
        ("sl2", "KO Stepladder", FINALE_ROUND_STEPLADDER, 3),
        ("f1", "KO-Finale", FINALE_ROUND_FINAL, 4),
        ("f2", "KO-Finale", FINALE_ROUND_FINAL, 5),
    ]
    if f3_col is not None:
        game_specs.append(("f3", "KO-Finale", FINALE_ROUND_FINAL, 6))

    for row_idx in range(3, (ws.max_row or 2) + 1):
        raw_name = ws.cell(row=row_idx, column=FINALE_COL_NAME).value
        if raw_name is None or not str(raw_name).strip():
            continue
        display_raw = str(raw_name).strip()
        display_name = _canonicalize_finale_name(display_raw)
        hdc = _as_int_score(ws.cell(row=row_idx, column=FINALE_COL_HDC).value)
        h_per_game = int(hdc) if hdc is not None else 0

        pid, _ = _resolve_player_id(display_name, player_lookup)
        if not pid:
            pid, _ = _resolve_player_id(display_raw, player_lookup)
        if not pid:
            unmatched.append(display_name)
            pid = f"UNK{hashlib.md5(display_name.encode('utf-8')).hexdigest()[:10]}"
        if pid in preferred_names_by_id:
            display_name = preferred_names_by_id[pid]

        for key, round_name, round_number, game_number in game_specs:
            if key == "f3":
                assert f3_col is not None
                col = f3_col
            else:
                col = FINALE_SCRATCH_COLS[key]
                if col is None:
                    continue
            score = _as_int_score(ws.cell(row=row_idx, column=col).value)
            if score is None:
                continue
            clean_rows.append(
                {
                    "Season": season,
                    "Date": event_date,
                    "Location": location,
                    "Event Type": "tournament",
                    "Event Name": EVENT_NAME,
                    "Round Number": str(round_number),
                    "Round Name": round_name,
                    "Player": display_name,
                    "Player ID": pid,
                    "Club": "",
                    "Game Number": str(game_number),
                    "Score": str(score),
                    "Handicap": str(h_per_game),
                }
            )

    return clean_rows, unmatched


def _is_lock_xlsx(path: Path) -> bool:
    name = path.name
    return name.startswith("~") or name.startswith(".~")


def _pick_qualifying_xlsx(input_dir: Path, explicit: str = "") -> Path:
    if explicit:
        return Path(explicit).resolve()
    preferred = input_dir / "Clubpokal DB 2026.xlsx"
    if preferred.is_file():
        return preferred.resolve()
    candidates = [
        p
        for p in input_dir.glob("*.xlsx")
        if not _is_lock_xlsx(p) and "finale" not in p.name.casefold()
    ]
    if not candidates:
        raise FileNotFoundError(f"No qualifying .xlsx files in {input_dir}")
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0].resolve()


def _pick_finale_xlsx(input_dir: Path, explicit: str = "") -> Optional[Path]:
    if explicit:
        p = Path(explicit).resolve()
        return p if p.is_file() else None
    candidates = [
        p
        for p in input_dir.glob("*.xlsx")
        if not _is_lock_xlsx(p) and "finale" in p.name.casefold()
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0].resolve()


def compute_workbook_fingerprint(
    xlsx_path: Path,
    *,
    season: str,
    event_date: str,
    location: str = "",
    sheet: str = "",
    first_data_row: int = 2,
) -> str:
    """SHA-256 of parsed sheet rows (same rules as import; ignores league Player ID lookup)."""
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = (sheet or "").strip() or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not in workbook. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    clean_rows, _ = _extract_rows_for_sheet(
        ws,
        season=season,
        event_date=event_date,
        location=location,
        player_lookup={},
        first_data_row=first_data_row,
    )
    if not clean_rows:
        raise ValueError(f"No score rows extracted from {xlsx_path}")

    tuples = [
        tuple(str(row.get(field, "") or "") for field in FINGERPRINT_FIELDS) for row in clean_rows
    ]
    tuples.sort()

    digest = hashlib.sha256()
    digest.update(f"{FINGERPRINT_VERSION}\0".encode())
    for row_tuple in tuples:
        digest.update("\x1e".join(row_tuple).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Import Clubmeisterschaft Donaubowler XLSX.")
    p.add_argument(
        "--xlsx",
        type=str,
        default="",
        help="Qualifying .xlsx path (default: Clubpokal DB 2026.xlsx or newest non-Finale in --input-dir).",
    )
    p.add_argument(
        "--finale-xlsx",
        type=str,
        default="",
        help="Finals .xlsx path (default: newest *Finale*.xlsx in --input-dir, if present).",
    )
    p.add_argument(
        "--no-finale",
        action="store_true",
        help="Skip finals workbook even if present.",
    )
    p.add_argument(
        "--input-dir",
        type=str,
        default=str(DEFAULT_INPUT_DIR),
        help="Directory containing Clubmeisterschaft workbook(s).",
    )
    p.add_argument(
        "--output",
        type=str,
        default=str(DEFAULT_OUTPUT),
        help="Batch postprocessed CSV output path.",
    )
    p.add_argument(
        "--league-csv",
        type=str,
        default=str(DEFAULT_LEAGUE_CSV),
        help="Logical league data path for Player ID lookup (CSV or Parquet sidecar).",
    )
    p.add_argument("--season", type=str, default="", help="Season label, e.g. 25/26 (default: from --year).")
    p.add_argument("--year", type=int, default=2026, help="Calendar year for default season label.")
    p.add_argument("--date", type=str, default="2026-05-15", help="ISO date for tournament rows.")
    p.add_argument("--location", type=str, default="", help="Venue / location string.")
    p.add_argument("--sheet", type=str, default="", help="Worksheet name (default: first sheet).")
    p.add_argument("--first-data-row", type=int, default=2, help="First Excel row index containing player data.")
    p.add_argument(
        "--fingerprint",
        action="store_true",
        help="Print workbook import fingerprint (parsed sheet data) and exit.",
    )
    return p


def main() -> None:
    args = build_parser().parse_args()
    input_dir = Path(args.input_dir).resolve()
    league_csv = Path(args.league_csv).resolve()

    xlsx_path = _pick_qualifying_xlsx(input_dir, str(args.xlsx or "").strip())
    finale_path = None if args.no_finale else _pick_finale_xlsx(input_dir, str(args.finale_xlsx or "").strip())

    if not xlsx_path.is_file():
        raise FileNotFoundError(xlsx_path)

    season = (args.season or "").strip() or _season_label_from_calendar_year(int(args.year))
    event_date = str(args.date).strip()
    location = str(args.location or "").strip()
    sheet_arg = str(args.sheet or "").strip()
    first_data_row = int(args.first_data_row)

    if args.fingerprint:
        print(
            compute_workbook_fingerprint(
                xlsx_path,
                season=season,
                event_date=event_date,
                location=location,
                sheet=sheet_arg,
                first_data_row=first_data_row,
            )
        )
        return

    bmi = _load_bayerische_import_module()
    player_lookup = _build_player_id_lookup(league_csv)

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = sheet_arg or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not in workbook. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    clean_rows, unmatched = _extract_rows_for_sheet(
        ws,
        season=season,
        event_date=event_date,
        location=location,
        player_lookup=player_lookup,
        first_data_row=first_data_row,
    )

    preferred_names_by_id: Dict[str, str] = {}
    for row in clean_rows:
        pid = str(row.get("Player ID") or "").strip()
        pname = str(row.get("Player") or "").strip()
        if pid and pname and pid not in preferred_names_by_id:
            preferred_names_by_id[pid] = pname

    finale_rows: List[Dict[str, str]] = []
    finale_unmatched: List[str] = []
    if finale_path and finale_path.is_file():
        fwb = load_workbook(finale_path, data_only=True)
        fws = fwb[fwb.sheetnames[0]]
        finale_rows, finale_unmatched = _extract_finale_rows_for_sheet(
            fws,
            season=season,
            event_date=event_date,
            location=location,
            player_lookup=player_lookup,
            preferred_names_by_id=preferred_names_by_id,
        )
        unmatched.extend(finale_unmatched)

    all_clean = clean_rows + finale_rows
    if not all_clean:
        raise ValueError(f"No score rows extracted from {xlsx_path}")

    postprocessed = bmi._postprocess(all_clean)
    output_path = Path(args.output).resolve()
    bmi._write_csv_rows(output_path, postprocessed)

    removed_from_gf = 0
    if GF_REGIONAL_TOURNAMENT_CSV.is_file():
        removed_from_gf = _strip_event_from_path(bmi, GF_REGIONAL_TOURNAMENT_CSV, EVENT_NAME)
    before_manual, after_manual = _merge_into_manual_postprocessed(bmi, postprocessed)
    _rebuild_player_hybrid_local()

    print(f"Source: {xlsx_path} (sheet {sheet_name!r})")
    if finale_path and finale_rows:
        print(f"Finale: {finale_path} ({len(finale_rows)} raw score rows)")
    print(f"Season: {season}  Event: {EVENT_NAME}")
    print(f"Wrote: {output_path} ({len(postprocessed)} postprocessed rows)")
    if removed_from_gf:
        print(
            f"Removed {removed_from_gf} rows for this event from GF export "
            f"({GF_REGIONAL_TOURNAMENT_CSV.name})"
        )
    print(
        f"Manual tournament CSV: {MANUAL_TOURNAMENT_CSV} "
        f"({before_manual} -> {after_manual} rows)"
    )
    print("Rebuilt player hybrid: database/data/player_stats_merged_plus_tournaments.csv")
    uniq_unmatched = sorted({n for n in unmatched if n})
    if uniq_unmatched:
        print(f"Warning: {len(uniq_unmatched)} player(s) without league ID match:")
        for n in uniq_unmatched[:30]:
            print(f"  - {n}")
        if len(uniq_unmatched) > 30:
            print(f"  ... and {len(uniq_unmatched) - 30} more")


if __name__ == "__main__":
    main()
